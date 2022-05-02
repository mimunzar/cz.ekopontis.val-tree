#!/usr/bin/env python3

import itertools as it
import functools as ft
import os
import openpyxl as xl
import openpyxl.styles as xl_styles
import openpyxl.utils  as xl_utils


def fit_row_height(ws, r_idx, n_chars):
    ws.row_dimensions[r_idx].height = 6*n_chars


def fit_col_width(ws, c_idx, n_chars):
    ws.column_dimensions[xl_utils.get_column_letter(c_idx)].width = 1.2*n_chars


def cell_alignment(params, c):
    c.alignment = xl_styles.Alignment(**params)
    return c


def cell_font(params, c):
    c.font = xl_styles.Font(**params)
    return c


def cell_comma_sep(c):
    c.number_format = '#,##'
    return c


DEFAULT_FONT = {'name': 'FreeMono', 'size': 8}

def cell(ws, r_idx, c_idx, x):
    return cell_font(DEFAULT_FONT, ws.cell(row=r_idx, column=c_idx, value=x))


def iter_cell_row(ws, r_idx, data_it):
    return it.starmap(ft.partial(cell, ws, r_idx), enumerate(data_it, 1))


def create_workbook():
    wb = xl.Workbook()
    wb.remove(wb.active)
    return wb


def load_workbook(fpath):
    return xl.load_workbook(fpath, data_only=True)


ALING_CENTER_2D = {'horizontal': 'center', 'vertical': 'center'}

def row_merged_cell(ws, r_idx, c_idx, n, x):
    ws.merge_cells(
        start_row    = r_idx,
        end_row      = r_idx + (n - 1),
        start_column = c_idx,
        end_column   = c_idx
    )
    return cell_alignment(ALING_CENTER_2D, cell(ws, r_idx, c_idx, x))


def col_merged_cell(ws, r_idx, c_idx, n, x):
    ws.merge_cells(
        start_row    = r_idx,
        end_row      = r_idx,
        start_column = c_idx,
        end_column   = c_idx + (n - 1)
    )
    return cell_alignment(ALING_CENTER_2D, cell(ws, r_idx, c_idx, x))


class ExcellAdapter:
    def __init__(self, fpath):
        self.fpath = fpath
        if os.path.exists(fpath):
            self.wb = load_workbook(fpath)
        else:
            self.wb = create_workbook()

    def open_sheet(self, title):
        if title in self.wb.get_sheet_names():
            return self.wb[title]
        return self.wb.create_sheet(title)

    def save(self):
        self.wb.save(filename=self.fpath)


def make(fpath):
    return ExcellAdapter(fpath)

