#!/usr/bin/env python3

import functools as ft
import sys
import openpyxl as xl
import requests

import src.val_tree.irecord as irecord
import src.val_tree.libs.util as util
import src.val_tree.libs.log as log


def send_request():
    data = {
        "taxon":"borovice černá (Pinus nigra)",
        "diameters":[1080],
        "diameters_on_stumps":[],
        "height":None,
        "stem_height":None,
        "spread":None,
        "vitality":"1",
        "health":"1",
        "removed_crown_volume":None,
        "location_attractiveness":"high",
        "growth_conditions":"unaffected",
        "microhabitats":[],
        "extensive_microhabitats":[],
        "taxon_offset":8,
        "_taxon_cz":"borovice černá",
        "_taxon_lat":"Pinus nigra",
        "memorial_tree":False,
        "deliberately_planted":False
    }

    r = requests.post(
            'https://ocenovanidrevin.nature.cz/hodnota-stromu.php',
            json=data,
            headers={'Content-Type': 'application/json'},
        )

def make_parser(n_rows):
    speed = log.make_avg_rec_sec()
    bar   = log.make_prog_bar(40, n_rows)
    empty = lambda c_it: all(map(lambda c: None == c, c_it))
    i     = 0
    print(f'{bar(i, 0)}', end='\r')
    def parser(acc, c_it):
        nonlocal i
        c_it = tuple(map(lambda c: c.value, c_it))
        if (not empty(c_it)):
            acc.append(irecord.parse(c_it))
        i = i + 1
        print(f'\033[K{bar(i, speed(1))}', end='\n' if n_rows == i else '\r')
        return acc
    return parser


if '__main__' == __name__:
    if 2 > len(sys.argv):
        print(log.fmt_err('Missing input file'), file=sys.stderr)
        sys.exit(1)

    wb = xl.load_workbook(filename=util.second(sys.argv), data_only=True)
    ws = wb.active

    print(log.fmt_msg('Parsing Sheet:\n'))
    result = ft.reduce(make_parser(ws.max_row - 1), util.drop(1, ws.iter_rows()), [])
    sys.exit(0)

