#!/usr/bin/env python3

import functools as ft
import openpyxl as xl

import src.val_tree.entities.measurement as measurement
import src.val_tree.entities.tree as tree
import src.val_tree.libs.log  as log
import src.val_tree.libs.util as util


def col_vals(c_it):
    return tuple(map(lambda c: c.value, c_it))


def empty_row(v_it):
    return all(map(lambda c: None == c, v_it))


def make_parser(n_rows):
    speed = log.make_avg_rec_sec()
    bar   = log.make_prog_bar(40, n_rows)
    i     = 0
    print(f'{bar(i, 0)}', end='\r')
    def parser(acc, val_it):
        nonlocal i
        i      = i + 1
        ma, ea = acc
        if (not empty_row(val_it)):
            m, e = measurement.parse_tree(val_it)
            if e: ea.append(log.fmt_tree_err(i + 1, e))
            else: ma.append(m)
        print(f'\033[K{bar(i, speed(1))}', end='\n\n' if n_rows == i else '\r')
        return acc
    return parser


def iter_tree_sheet(fpath):
    ws = xl.load_workbook(filename=fpath, data_only=True).active
    return ft.reduce(make_parser(ws.max_row - 1),
            map(col_vals, util.drop(1, ws.iter_rows())), [[], []])

